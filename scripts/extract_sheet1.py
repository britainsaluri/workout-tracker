#!/usr/bin/env python3
"""
Extract workout data from Sheet 1 ("Davey Jone's Pump") of workout Excel file.
Handles complex multi-week, multi-day layout with exercise details.
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def clean_cell_value(value: Any) -> Optional[str]:
    """Clean and normalize cell values."""
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str if value_str else None


def is_day_header(row_values: List[Any]) -> Optional[str]:
    """Check if row is a day header (e.g., 'DAY 1: Upper Pull/Lower Push')."""
    first_cell = clean_cell_value(row_values[0])
    if first_cell and first_cell.upper().startswith('DAY '):
        return first_cell
    return None


def extract_day_number(day_header: str) -> int:
    """Extract day number from header like 'DAY 1: Upper Pull/Lower Push'."""
    try:
        # Extract number after 'DAY '
        parts = day_header.split(':')[0].strip()
        day_num = ''.join(filter(str.isdigit, parts))
        return int(day_num) if day_num else 0
    except:
        return 0


def parse_exercise_row(row_values: List[Any], col_offset: int = 0) -> Optional[Dict[str, str]]:
    """
    Parse exercise row with structure:
    [ID, Exercise Name, Tempo, Sets/Reps, Rest, (optional) Results]

    Args:
        row_values: List of cell values
        col_offset: Column offset for Week 2 (if applicable)
    """
    # Adjust for column offset
    idx_id = col_offset
    idx_name = col_offset + 1
    idx_tempo = col_offset + 2
    idx_sets = col_offset + 3
    idx_rest = col_offset + 4
    idx_results = col_offset + 5 if len(row_values) > col_offset + 5 else None

    exercise_id = clean_cell_value(row_values[idx_id]) if idx_id < len(row_values) else None
    exercise_name = clean_cell_value(row_values[idx_name]) if idx_name < len(row_values) else None

    # Check if this is a valid exercise row (has ID like A1, B1, etc.)
    if not exercise_id or not exercise_name:
        return None

    # Validate ID format (should be letter + number)
    if not any(c.isalpha() for c in exercise_id) or not any(c.isdigit() for c in exercise_id):
        return None

    tempo = clean_cell_value(row_values[idx_tempo]) if idx_tempo < len(row_values) else ""
    sets_reps = clean_cell_value(row_values[idx_sets]) if idx_sets < len(row_values) else ""
    rest = clean_cell_value(row_values[idx_rest]) if idx_rest < len(row_values) else ""
    results = clean_cell_value(row_values[idx_results]) if idx_results and idx_results < len(row_values) else ""

    return {
        "id": exercise_id,
        "name": exercise_name,
        "tempo": tempo or "",
        "setsReps": sets_reps or "",
        "rest": rest or "",
        "results": results or ""
    }


def extract_week_data(sheet: Any, week_num: int, col_offset: int = 0) -> Dict[str, Any]:
    """
    Extract data for a specific week.

    Args:
        sheet: openpyxl worksheet object
        week_num: Week number (1 or 2)
        col_offset: Column offset (0 for Week 1, ~6-7 for Week 2)
    """
    week_data = {
        "week": week_num,
        "days": []
    }

    current_day = None
    current_exercises = []

    # Iterate through rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not any(row):  # Skip empty rows
            continue

        # Check if this is a day header
        day_header = is_day_header(row)
        if day_header:
            # Save previous day if exists
            if current_day and current_exercises:
                current_day["exercises"] = current_exercises
                week_data["days"].append(current_day)

            # Start new day
            current_day = {
                "dayNumber": extract_day_number(day_header),
                "dayName": day_header,
                "exercises": []
            }
            current_exercises = []
            continue

        # Try to parse as exercise row
        if current_day:
            exercise = parse_exercise_row(row, col_offset)
            if exercise:
                current_exercises.append(exercise)

    # Save last day if exists
    if current_day and current_exercises:
        current_day["exercises"] = current_exercises
        week_data["days"].append(current_day)

    return week_data


def extract_sheet1_data(excel_path: str) -> Dict[str, Any]:
    """
    Extract workout data from Sheet 1.

    Args:
        excel_path: Path to Excel file

    Returns:
        Dictionary with structured workout data
    """
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path, data_only=True)

        # Get first sheet (Sheet 1)
        sheet = workbook.worksheets[0]
        sheet_name = sheet.title

        print(f"Processing sheet: {sheet_name}")
        print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        # Initialize output structure
        output = {
            "program": sheet_name,
            "weeks": []
        }

        # Extract Week 1 data (columns 0-5 typically)
        print("\nExtracting Week 1 data...")
        week1_data = extract_week_data(sheet, week_num=1, col_offset=0)
        if week1_data["days"]:
            output["weeks"].append(week1_data)
            print(f"  Found {len(week1_data['days'])} days")

        # Extract Week 2 data (columns 6-11 typically)
        # Try to detect Week 2 column offset by looking for "WEEK 2" header
        week2_col_offset = None
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            for col_idx, cell in enumerate(row):
                if cell and "WEEK 2" in str(cell).upper():
                    week2_col_offset = col_idx
                    break
            if week2_col_offset:
                break

        if week2_col_offset:
            print(f"\nExtracting Week 2 data (column offset: {week2_col_offset})...")
            week2_data = extract_week_data(sheet, week_num=2, col_offset=week2_col_offset)
            if week2_data["days"]:
                output["weeks"].append(week2_data)
                print(f"  Found {len(week2_data['days'])} days")

        workbook.close()

        # Summary
        total_days = sum(len(week["days"]) for week in output["weeks"])
        total_exercises = sum(
            len(day["exercises"])
            for week in output["weeks"]
            for day in week["days"]
        )

        print(f"\n✓ Extraction complete!")
        print(f"  Program: {output['program']}")
        print(f"  Weeks: {len(output['weeks'])}")
        print(f"  Total Days: {total_days}")
        print(f"  Total Exercises: {total_exercises}")

        return output

    except FileNotFoundError:
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    """Main execution function."""
    # File paths
    excel_path = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    output_path = "/Users/britainsaluri/workout-tracker/src/sheet1-workout-data.json"

    print("=" * 60)
    print("Sheet 1 Workout Data Extractor")
    print("=" * 60)
    print(f"\nInput:  {excel_path}")
    print(f"Output: {output_path}\n")

    # Check if input file exists
    if not Path(excel_path).exists():
        print(f"❌ Error: Input file not found: {excel_path}")
        sys.exit(1)

    # Extract data
    workout_data = extract_sheet1_data(excel_path)

    # Save to JSON
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(workout_data, f, indent=2, ensure_ascii=False)

    print(f"\n✓ Data saved to: {output_path}")
    print(f"  File size: {Path(output_path).stat().st_size:,} bytes")

    return 0


if __name__ == "__main__":
    sys.exit(main())
