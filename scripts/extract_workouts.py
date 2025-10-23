#!/usr/bin/env python3
"""
Workout Data Extraction Script
Extracts workout data from Excel file and converts to structured JSON format.
Handles multiple sheets with different layouts (Sheet1-4).
"""

import openpyxl
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
import re


class WorkoutExtractor:
    """Extract and structure workout data from Excel workbook."""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)
        self.workout_data = {
            "program_name": "Argh Let's Get Huge Matey",
            "sheets": []
        }

    def extract_all_sheets(self) -> Dict[str, Any]:
        """Extract data from all sheets in the workbook."""
        for sheet_name in self.workbook.sheetnames[:4]:  # Process Sheet1-4
            print(f"Processing {sheet_name}...")
            ws = self.workbook[sheet_name]

            if sheet_name == "Sheet4":
                sheet_data = self._extract_sheet4(ws, sheet_name)
            else:
                sheet_data = self._extract_standard_sheet(ws, sheet_name)

            if sheet_data:
                self.workout_data["sheets"].append(sheet_data)

        return self.workout_data

    def _extract_standard_sheet(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Extract data from Sheet1, Sheet2, Sheet3 (standard format)."""
        rows = list(ws.iter_rows(values_only=True))

        # Get program name from first row
        program_name = None
        for cell in rows[0]:
            if cell and isinstance(cell, str) and cell.strip():
                program_name = cell.strip()
                break

        sheet_data = {
            "sheet_name": sheet_name,
            "program_name": program_name,
            "days": []
        }

        current_day = None
        week_headers = []

        for i, row in enumerate(rows):
            # Detect week headers (e.g., WEEK 1, WEEK 2, etc.)
            if row[3] and isinstance(row[3], str) and "WEEK" in str(row[3]).upper():
                week_headers = self._extract_week_headers(row)
                continue

            # Detect day header
            if row[0] and isinstance(row[0], str) and "DAY" in str(row[0]).upper():
                if current_day:
                    sheet_data["days"].append(current_day)

                current_day = {
                    "day_name": str(row[0]).strip(),
                    "weeks": self._initialize_weeks(week_headers)
                }
                continue

            # Extract exercise data
            if current_day and row[0] and isinstance(row[0], str):
                exercise_id = str(row[0]).strip()
                if re.match(r'^[A-Z]\d+$', exercise_id):  # Match A1, B2, etc.
                    self._add_exercise_to_day(current_day, row, week_headers)

        # Add last day
        if current_day:
            sheet_data["days"].append(current_day)

        return sheet_data

    def _extract_sheet4(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Extract data from Sheet4 (different format)."""
        rows = list(ws.iter_rows(values_only=True))

        sheet_data = {
            "sheet_name": sheet_name,
            "program_name": "Britanica",
            "days": []
        }

        current_day = None
        current_block = None
        week_range = None

        for i, row in enumerate(rows):
            # Detect week range
            if row[0] and isinstance(row[0], str) and "Week" in str(row[0]):
                week_range = str(row[0]).strip()
                continue

            # Detect day header
            if row[0] and isinstance(row[0], str) and "Day" in str(row[0]):
                if current_day:
                    sheet_data["days"].append(current_day)

                current_day = {
                    "day_name": str(row[0]).strip(),
                    "week_range": week_range,
                    "blocks": []
                }
                continue

            # Detect block header
            if row[0] and isinstance(row[0], str) and "Block" in str(row[0]):
                current_block = {
                    "block_name": str(row[0]).strip(),
                    "exercises": []
                }
                if current_day:
                    current_day["blocks"].append(current_block)
                continue

            # Extract exercise data for Sheet4
            if current_block and row[0] and isinstance(row[0], str) and not any(x in str(row[0]) for x in ["Block", "Day", "Week"]):
                exercise = self._parse_sheet4_exercise(row)
                if exercise:
                    current_block["exercises"].append(exercise)

        # Add last day
        if current_day:
            sheet_data["days"].append(current_day)

        return sheet_data

    def _extract_week_headers(self, row: tuple) -> List[Dict[str, Any]]:
        """Extract week information from header row."""
        weeks = []
        week_positions = {}

        for i, cell in enumerate(row):
            if cell and isinstance(cell, str) and "WEEK" in str(cell).upper():
                week_num = re.search(r'\d+', str(cell))
                if week_num:
                    weeks.append({
                        "week_number": int(week_num.group()),
                        "column_start": i
                    })

        return weeks

    def _initialize_weeks(self, week_headers: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Initialize week structure for a day."""
        return [{
            "week_number": week["week_number"],
            "exercises": []
        } for week in week_headers]

    def _add_exercise_to_day(self, current_day: Dict[str, Any], row: tuple, week_headers: List[Dict[str, Any]]):
        """Add exercise data to current day for each week."""
        exercise_id = str(row[0]).strip()
        exercise_name = str(row[1]).strip() if row[1] else ""

        # For each week, extract exercise data
        for week_idx, week in enumerate(week_headers):
            # Calculate column offset for this week
            col_offset = week["column_start"]

            tempo = self._safe_get_value(row, col_offset)
            sets_reps = self._safe_get_value(row, col_offset + 1)
            rest = self._safe_get_value(row, col_offset + 2)
            results = self._safe_get_value(row, col_offset + 3)

            exercise = {
                "exercise_id": exercise_id,
                "exercise_name": exercise_name,
                "tempo": self._format_tempo(tempo),
                "sets_reps": str(sets_reps) if sets_reps else "",
                "rest": str(rest) if rest else "",
                "results": str(results) if results else ""
            }

            # Add to appropriate week
            if week_idx < len(current_day["weeks"]):
                current_day["weeks"][week_idx]["exercises"].append(exercise)

    def _parse_sheet4_exercise(self, row: tuple) -> Optional[Dict[str, Any]]:
        """Parse exercise data from Sheet4 format."""
        exercise_name = str(row[0]).strip() if row[0] else ""

        if not exercise_name or len(exercise_name) < 3:
            return None

        return {
            "exercise_name": exercise_name,
            "week_1_2": {
                "tempo": self._format_tempo(self._safe_get_value(row, 2)),
                "sets": str(self._safe_get_value(row, 3)) if row[3] else "",
                "reps": str(self._safe_get_value(row, 4)) if row[4] else "",
                "results": str(self._safe_get_value(row, 5)) if row[5] else "",
                "pump_rating": self._safe_get_value(row, 7)
            },
            "week_3_4": {
                "tempo": self._format_tempo(self._safe_get_value(row, 9)),
                "sets": str(self._safe_get_value(row, 10)) if row[10] else "",
                "reps": str(self._safe_get_value(row, 11)) if row[11] else "",
                "results": str(self._safe_get_value(row, 12)) if row[12] else "",
                "pump_rating": self._safe_get_value(row, 14)
            }
        }

    def _safe_get_value(self, row: tuple, index: int) -> Any:
        """Safely get value from row at index."""
        try:
            if index < len(row):
                value = row[index]
                # Handle datetime objects
                if isinstance(value, datetime):
                    return value.strftime("%m-%d")
                return value
            return None
        except IndexError:
            return None

    def _format_tempo(self, tempo: Any) -> str:
        """Format tempo value consistently."""
        if tempo is None:
            return ""
        if isinstance(tempo, (int, float)):
            # Convert 211.0 to "211", 311.0 to "311", etc.
            return str(int(tempo))
        return str(tempo)

    def save_to_json(self, output_path: str):
        """Save extracted data to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.workout_data, f, indent=2, ensure_ascii=False)
        print(f"✅ Data saved to {output_path}")

    def close(self):
        """Close the workbook."""
        self.workbook.close()


def main():
    """Main execution function."""
    # Paths
    excel_path = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    output_path = "/Users/britainsaluri/workout-tracker/src/workout-data.json"

    print("🏋️  Starting workout data extraction...")
    print(f"📂 Input: {excel_path}")
    print(f"📝 Output: {output_path}")
    print("-" * 80)

    # Extract data
    extractor = WorkoutExtractor(excel_path)
    workout_data = extractor.extract_all_sheets()

    # Save to JSON
    extractor.save_to_json(output_path)

    # Print summary
    print("\n" + "=" * 80)
    print("📊 EXTRACTION SUMMARY")
    print("=" * 80)
    print(f"Total sheets processed: {len(workout_data['sheets'])}")

    for sheet in workout_data['sheets']:
        print(f"\n{sheet['sheet_name']} - {sheet.get('program_name', 'N/A')}")
        print(f"  Days: {len(sheet['days'])}")

        for day in sheet['days']:
            print(f"    - {day['day_name']}")
            if 'weeks' in day:
                for week in day['weeks']:
                    print(f"      Week {week['week_number']}: {len(week['exercises'])} exercises")
            elif 'blocks' in day:
                for block in day['blocks']:
                    print(f"      {block['block_name']}: {len(block['exercises'])} exercises")

    extractor.close()
    print("\n✅ Extraction complete!")


if __name__ == "__main__":
    main()
