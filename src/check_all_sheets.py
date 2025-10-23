#!/usr/bin/env python3
"""
Check all sheets in the Excel workbook to locate Week 2 data
"""

import openpyxl
import sys

def check_workbook_sheets(file_path):
    """List all sheets and preview their content"""

    print(f"Loading workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    print(f"\nTotal sheets in workbook: {len(workbook.worksheets)}")
    print("=" * 80)

    for idx, sheet in enumerate(workbook.worksheets, start=1):
        print(f"\n### Sheet {idx}: {sheet.title} ###")
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        # Preview first 20 rows
        print("\nFirst 20 rows preview (Column A-F):")
        print("-" * 80)

        for row_num in range(1, min(21, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(7, sheet.max_column + 1)):
                cell = sheet.cell(row_num, col)
                value = str(cell.value).strip() if cell.value else ""
                if len(value) > 30:
                    value = value[:27] + "..."
                row_data.append(value)

            # Only print non-empty rows
            if any(row_data):
                print(f"Row {row_num:3d}: {' | '.join(row_data)}")

        # Check for Week 2 indicators
        print("\n--- Checking for Week 2 indicators ---")
        week2_found = False
        for row_num in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell_value = str(sheet.cell(row_num, col).value or "").upper()
                if "WEEK 2" in cell_value or "WEEK2" in cell_value:
                    print(f"✓ Found 'Week 2' at Row {row_num}, Column {col}")
                    week2_found = True

        if not week2_found:
            print("✗ No 'Week 2' indicators found in this sheet")

        print("=" * 80)

    workbook.close()

if __name__ == "__main__":
    input_file = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    try:
        check_workbook_sheets(input_file)
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
