#!/usr/bin/env python3
"""
Inspect exact location of Week 2 data
"""

import openpyxl
import sys

def inspect_week2(file_path):
    """Inspect Week 2 data location"""

    print(f"Loading workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    sheet = workbook.worksheets[0]
    print(f"\nSheet name: {sheet.title}")
    print(f"Max rows: {sheet.max_row}")
    print(f"Max columns: {sheet.max_column}")

    print("\n=== Row 2 (Week headers) - All columns ===\n")
    for col in range(1, 17):
        cell = sheet.cell(2, col)
        value = str(cell.value).strip() if cell.value else ""
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"Column {col_letter} (index {col}): '{value}'")

    print("\n=== Rows 3-6 - All columns (Day 1 data) ===\n")
    for row_num in range(3, 7):
        row_data = []
        for col in range(1, 17):
            cell = sheet.cell(row_num, col)
            value = str(cell.value).strip() if cell.value else ""
            if len(value) > 20:
                value = value[:17] + "..."
            row_data.append(value)

        print(f"Row {row_num}:")
        for idx, val in enumerate(row_data, start=1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            if val:
                print(f"  {col_letter}: {val}")

    workbook.close()

if __name__ == "__main__":
    input_file = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    try:
        inspect_week2(input_file)
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
