#!/usr/bin/env python3
"""
Extract COMPLETE workout data from Sheet 1 (Davey Jone's Pump)
Extracts all 5 workout days for BOTH Week 1 AND Week 2
Week 1 is in columns A-F, Week 2 is in columns H-M (starting at column 8)
"""

import openpyxl
import json
import sys
from pathlib import Path

def clean_cell_value(cell):
    """Extract and clean cell value"""
    if cell is None or cell.value is None:
        return ""
    value = str(cell.value).strip()
    # Remove .0 from tempo values like "211.0" -> "211"
    if value.replace('.', '').replace('0', '').isdigit() and value.endswith('.0'):
        value = value.replace('.0', '')
    return value

def parse_exercise_row(row_cells, day_num, week_num, week_config):
    """Parse a single exercise row
    Week 1: ID(A=0), Name(B=1), Tempo(D=3), Sets(E=4), Rest(F=5), Results(G=6)
    Week 2: ID(A=0), Name(B=1), Tempo(I=8), Sets(J=9), Rest(K=10), Results(L=11)
    """
    # Exercise ID and Name are always in columns A and B
    exercise_id = clean_cell_value(row_cells[0])
    exercise_name = clean_cell_value(row_cells[1])

    # Tempo, Sets/Reps, Rest, Results are in different columns per week
    tempo_col = week_config["tempo_col"]
    sets_col = week_config["sets_col"]
    rest_col = week_config["rest_col"]
    results_col = week_config["results_col"]

    exercise = {
        "exercise_id": exercise_id,
        "exercise_name": exercise_name,
        "tempo": clean_cell_value(row_cells[tempo_col]),
        "sets_reps": clean_cell_value(row_cells[sets_col]),
        "rest": clean_cell_value(row_cells[rest_col]),
        "results": clean_cell_value(row_cells[results_col]),
        "day": day_num,
        "week": week_num
    }
    return exercise

def extract_sheet1_data(file_path):
    """Extract all workout data from Sheet 1 - Both Week 1 and Week 2"""

    print(f"Loading workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Get Sheet 1
    sheet = workbook.worksheets[0]
    print(f"Processing sheet: {sheet.title}")

    workout_data = {
        "sheet_name": sheet.title,
        "program_name": "Davey Jone's Pump",
        "weeks": []
    }

    max_rows = sheet.max_row
    print(f"Total rows in sheet: {max_rows}")

    # Define day boundaries (same for both weeks)
    day_configs = [
        {"name": "DAY 1", "start": 3, "end": 13},   # Rows 3-13
        {"name": "DAY 2", "start": 15, "end": 22},  # Rows 15-22
        {"name": "DAY 3", "start": 24, "end": 29},  # Rows 24-29
        {"name": "DAY 4", "start": 31, "end": 38},  # Rows 31-38
        {"name": "DAY 5", "start": 40, "end": 50},  # Rows 40-50
    ]

    # Process both weeks
    # Week 1: ID(A), Name(B), Tempo(D), Sets(E), Rest(F), Results(G)
    # Week 2: ID(A), Name(B), Tempo(I), Sets(J), Rest(K), Results(L)
    week_configs = [
        {
            "week": 1,
            "tempo_col": 3,   # Column D
            "sets_col": 4,    # Column E
            "rest_col": 5,    # Column F
            "results_col": 6, # Column G
            "label": "Week 1 (Columns A,B,D,E,F,G)"
        },
        {
            "week": 2,
            "tempo_col": 8,   # Column I
            "sets_col": 9,    # Column J
            "rest_col": 10,   # Column K
            "results_col": 11,# Column L
            "label": "Week 2 (Columns A,B,I,J,K,L)"
        }
    ]

    for week_config in week_configs:
        week_num = week_config["week"]

        week_data = {
            "week": week_num,
            "days": []
        }

        print(f"\n{'=' * 80}")
        print(f"=== Processing {week_config['label']} ===")
        print(f"{'=' * 80}")

        for day_num, config in enumerate(day_configs, start=1):
            day_data = {
                "day": day_num,
                "day_name": "",
                "exercises": []
            }

            print(f"\n--- Day {day_num} (rows {config['start']}-{config['end']}) ---")

            # Get day name from first row (always in column A)
            if week_num == 1:
                day_header = clean_cell_value(sheet.cell(config['start'], 1))
                if ":" in day_header:
                    day_data["day_name"] = day_header.split(":", 1)[1].strip()
                    print(f"Day name: {day_data['day_name']}")

            # Skip header row, start with exercises
            exercise_count = 0
            for row_num in range(config['start'] + 1, config['end'] + 1):
                row_cells = []
                # Get all columns A-L (indices 0-11)
                for col in range(1, 13):
                    row_cells.append(sheet.cell(row_num, col))

                # Exercise ID and Name are always in columns A and B
                exercise_id = clean_cell_value(row_cells[0])
                exercise_name = clean_cell_value(row_cells[1])

                # Skip empty rows
                if not exercise_id or not exercise_name:
                    continue

                # Skip if it's a header row
                if "TEMPO" in exercise_id.upper() or "SETS" in exercise_id.upper():
                    continue

                # Valid exercise row - parse using week-specific column mappings
                exercise = parse_exercise_row(row_cells, day_num, week_num, week_config)
                day_data["exercises"].append(exercise)
                exercise_count += 1
                print(f"  {exercise_count:2d}. {exercise['exercise_id']:3s} - {exercise['exercise_name'][:50]:<50s} | Tempo: {exercise['tempo']:>4s} | Sets: {exercise['sets_reps']:<12s} | Rest: {exercise['rest']}")

            # Add day name from Week 1 if processing Week 2
            if week_num == 2 and not day_data["day_name"]:
                # Find day name from the corresponding week 1 day
                for week in workout_data["weeks"]:
                    if week["week"] == 1:
                        for day in week["days"]:
                            if day["day"] == day_num:
                                day_data["day_name"] = day["day_name"]
                                break

            print(f"Day {day_num} total exercises: {exercise_count}")
            week_data["days"].append(day_data)

        workout_data["weeks"].append(week_data)
        print(f"\nWeek {week_num} complete")

    workbook.close()
    return workout_data

def main():
    input_file = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    output_file = "/Users/britainsaluri/workout-tracker/src/sheet1-workout-data.json"

    try:
        # Extract data
        workout_data = extract_sheet1_data(input_file)

        # Save to JSON
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(workout_data, f, indent=2, ensure_ascii=False)

        print(f"\n{'=' * 80}")
        print(f"✓ Data extracted successfully!")
        print(f"✓ Output saved to: {output_file}")
        print(f"{'=' * 80}")

        # Print summary
        print("\n=== EXTRACTION SUMMARY ===")
        total_exercises = 0
        for week in workout_data["weeks"]:
            print(f"\nWeek {week['week']}:")
            week_total = 0
            for day in week["days"]:
                exercise_count = len(day["exercises"])
                week_total += exercise_count
                print(f"  Day {day['day']} ({day['day_name']}): {exercise_count} exercises")
            print(f"  Week {week['week']} Total: {week_total} exercises")
            total_exercises += week_total

        print(f"\n📊 GRAND TOTAL: {total_exercises} exercises across {len(workout_data['weeks'])} weeks")

        # Verify counts for both weeks
        print("\n=== VERIFICATION ===")
        expected_counts = {1: 10, 2: 7, 3: 5, 4: 7, 5: 10}
        all_correct = True

        for week in workout_data["weeks"]:
            print(f"\nWeek {week['week']}:")
            for day in week["days"]:
                actual = len(day["exercises"])
                expected = expected_counts.get(day["day"], "?")
                status = "✓" if actual == expected else "✗"
                if actual != expected:
                    all_correct = False
                print(f"  {status} Day {day['day']}: Expected {expected}, Got {actual}")

        if all_correct:
            print("\n✅ All exercise counts match expected values for both weeks!")
        else:
            print("\n⚠️  Some exercise counts don't match expected values")

        return 0

    except FileNotFoundError:
        print(f"ERROR: File not found: {input_file}")
        return 1
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
