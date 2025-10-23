#!/usr/bin/env python3
"""
Inspect Sheet 1 structure to understand layout
"""

import openpyxl
import sys

def inspect_sheet(file_path):
    """Inspect sheet structure"""

    print(f"Loading workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Get Sheet 1
    sheet = workbook.worksheets[0]
    print(f"\nSheet name: {sheet.title}")
    print(f"Max rows: {sheet.max_row}")
    print(f"Max columns: {sheet.max_column}")

    print("\n=== First 100 rows (Column A-F) ===\n")

    for row_num in range(1, min(101, sheet.max_row + 1)):
        row_data = []
        for col in range(1, 7):
            cell = sheet.cell(row_num, col)
            value = str(cell.value).strip() if cell.value else ""
            row_data.append(value)

        # Only print non-empty rows
        if any(row_data):
            print(f"Row {row_num:3d}: {' | '.join(row_data)}")

    workbook.close()

if __name__ == "__main__":
    input_file = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    try:
        inspect_sheet(input_file)
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
