#!/usr/bin/env python3
"""
Extract complete workout data from Sheet 1 (Davey Jone's Pump)
Extracts all 5 workout days for Week 1 (rows 3-50 only show Week 1)
Note: Based on inspection, only Week 1 data is present in Sheet 1
"""

import openpyxl
import json
import sys
from pathlib import Path

def clean_cell_value(cell):
    """Extract and clean cell value"""
    if cell is None or cell.value is None:
        return ""
    value = str(cell.value).strip()
    # Remove .0 from tempo values like "211.0" -> "211"
    if value.replace('.', '').replace('0', '').isdigit() and value.endswith('.0'):
        value = value.replace('.0', '')
    return value

def parse_exercise_row(row_cells, day_num, week_num):
    """Parse a single exercise row"""
    exercise = {
        "exercise_id": clean_cell_value(row_cells[0]),
        "exercise_name": clean_cell_value(row_cells[1]),
        "tempo": clean_cell_value(row_cells[3]),  # Column D (index 3)
        "sets_reps": clean_cell_value(row_cells[4]),  # Column E (index 4)
        "rest": clean_cell_value(row_cells[5]),  # Column F (index 5)
        "day": day_num,
        "week": week_num
    }
    return exercise

def extract_sheet1_data(file_path):
    """Extract all workout data from Sheet 1"""

    print(f"Loading workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Get Sheet 1
    sheet = workbook.worksheets[0]
    print(f"Processing sheet: {sheet.title}")

    workout_data = {
        "sheet_name": sheet.title,
        "program_name": "Davey Jone's Pump",
        "weeks": []
    }

    max_rows = sheet.max_row
    print(f"Total rows in sheet: {max_rows}")

    # Process Week 1 (only week present in Sheet 1)
    week_data = {
        "week": 1,
        "days": []
    }

    print(f"\n=== Processing Week 1 ===")

    # Define day boundaries based on inspection
    day_configs = [
        {"name": "DAY 1", "start": 3, "end": 13},   # Rows 3-13
        {"name": "DAY 2", "start": 15, "end": 22},  # Rows 15-22 (actually Day 2)
        {"name": "DAY 3", "start": 24, "end": 29},  # Rows 24-29
        {"name": "DAY 4", "start": 31, "end": 38},  # Rows 31-38
        {"name": "DAY 5", "start": 40, "end": 50},  # Rows 40-50
    ]

    for day_num, config in enumerate(day_configs, start=1):
        day_data = {
            "day": day_num,
            "day_name": "",
            "exercises": []
        }

        print(f"\n--- Processing Day {day_num} (rows {config['start']}-{config['end']}) ---")

        # Get day name from first row
        day_header = clean_cell_value(sheet.cell(config['start'], 1))
        if ":" in day_header:
            day_data["day_name"] = day_header.split(":", 1)[1].strip()
            print(f"Day name: {day_data['day_name']}")

        # Skip header row, start with exercises
        exercise_count = 0
        for row_num in range(config['start'] + 1, config['end'] + 1):
            row_cells = []
            for col in range(1, 7):
                row_cells.append(sheet.cell(row_num, col))

            # Check if this is an exercise row (has ID in first column)
            exercise_id = clean_cell_value(row_cells[0])
            exercise_name = clean_cell_value(row_cells[1])

            # Skip empty rows
            if not exercise_id or not exercise_name:
                continue

            # Skip if it's a header row
            if "TEMPO" in exercise_id.upper() or "SETS" in exercise_id.upper():
                continue

            # Valid exercise row
            exercise = parse_exercise_row(row_cells, day_num, 1)
            day_data["exercises"].append(exercise)
            exercise_count += 1
            print(f"  {exercise_count:2d}. {exercise['exercise_id']:3s} - {exercise['exercise_name'][:50]:<50s} | Tempo: {exercise['tempo']:>4s} | Sets: {exercise['sets_reps']:<12s} | Rest: {exercise['rest']}")

        print(f"Day {day_num} total exercises: {exercise_count}")
        week_data["days"].append(day_data)

    workout_data["weeks"].append(week_data)
    print(f"\nWeek 1 complete")

    workbook.close()
    return workout_data

def main():
    input_file = "/Users/britainsaluri/Downloads/Argh Let's Get Huge Matey.xlsx"
    output_file = "/Users/britainsaluri/workout-tracker/src/sheet1-workout-data.json"

    try:
        # Extract data
        workout_data = extract_sheet1_data(input_file)

        # Save to JSON
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(workout_data, f, indent=2, ensure_ascii=False)

        print(f"\n✓ Data extracted successfully!")
        print(f"✓ Output saved to: {output_file}")

        # Print summary
        print("\n=== EXTRACTION SUMMARY ===")
        for week in workout_data["weeks"]:
            print(f"\nWeek {week['week']}:")
            for day in week["days"]:
                exercise_count = len(day["exercises"])
                print(f"  Day {day['day']} ({day['day_name']}): {exercise_count} exercises")

        # Verify counts
        print("\n=== VERIFICATION ===")
        expected_counts = {1: 10, 2: 7, 3: 5, 4: 7, 5: 10}
        all_correct = True
        for week in workout_data["weeks"]:
            for day in week["days"]:
                actual = len(day["exercises"])
                expected = expected_counts.get(day["day"], "?")
                status = "✓" if actual == expected else "✗"
                if actual != expected:
                    all_correct = False
                print(f"  {status} Day {day['day']}: Expected {expected}, Got {actual}")

        if all_correct:
            print("\n✓ All exercise counts match expected values!")
        else:
            print("\n⚠ Some exercise counts don't match expected values")

        return 0

    except FileNotFoundError:
        print(f"ERROR: File not found: {input_file}")
        return 1
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
